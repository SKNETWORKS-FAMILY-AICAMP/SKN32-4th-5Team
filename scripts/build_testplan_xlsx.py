"""docs/13_테스트계획.md → 제출용 xlsx.

    python scripts/build_testplan_xlsx.py --write   # 제출/테스트계획및결과보고서.xlsx
    python scripts/build_testplan_xlsx.py           # 미리 보기만
    make testplan-xlsx                              # 같은 것 (Windows 에는 make 가 없다)

🔴 **xlsx 는 생성물이다. 손으로 고치지 않는다.** `build_requirements_xlsx.py` 와 같은 규약이다.

사람이 고치는 것은 `docs/13` 하나뿐이다. 팀원 원본(`docs/ksr/…xlsx`)은 **기여 근거로**
남고, 거기서 정본으로 옮겨 적는 일은 `docs/13 §2.2` 에서 한 번만 일어난다.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "13_테스트계획.md"
OUT = ROOT / "제출" / "테스트계획및결과보고서.xlsx"

#: §2.2 의 테스트케이스 행. `| \`AUTH-001\` | FR-01 · FR-02 | ✅ 통과 | 관측 |`
_TC_ROW = re.compile(r"^\|\s*`([A-Z0-9-]+)`\s*\|([^|]*)\|([^|]*)\|(.*)\|\s*$")

#: §1 의 층 요약. `① 단위     603건  프레임워크 무관 — 그대로 승계        §3`
_LAYER = re.compile(
    # 층 이름에 공백이 있다 (`수동 UI`) — `\S+` 로는 못 읽는다.
    r"^([①②③④⑤])\s*(.+?)\s+(\d+)건\s+(.*?)\s+(§[\d.]+)\s*$"
)


def _plain(cell: str) -> str:
    cell = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", cell)
    # 🔴 **밑줄 이탤릭은 다루지 않는다.** `_..._` 를 벗기려 들면 `test_auth_api.py` 의
    #    밑줄까지 먹고, 경계 조건을 붙여도 `_@login_required_` 처럼 **내용에 밑줄이 있는**
    #    경우를 못 푼다. 그래서 §2.2 표는 이탤릭에 별표를 쓴다 — 아래 한 줄이 처리한다.
    return re.sub(r"[*`~]", "", cell).strip()


def read_cases() -> list[list[str]]:
    rows: list[list[str]] = []
    for line in DOC.read_text(encoding="utf-8").splitlines():
        m = _TC_ROW.match(line.strip())
        if m:
            rows.append([_plain(g) for g in m.groups()])
    return rows


def read_layers() -> list[list[str]]:
    rows: list[list[str]] = []
    for line in DOC.read_text(encoding="utf-8").splitlines():
        m = _LAYER.match(line.strip())
        if m:
            rows.append([m.group(1), m.group(2), m.group(3), _plain(m.group(4)), m.group(5)])
    return rows


def _write(layers: list[list[str]], cases: list[list[str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl 이 없다 — pip install -e '.[ingest]' -c constraints.txt")

    head_fill = PatternFill("solid", fgColor="3A4F37")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    sheets = [
        ("테스트 층", ["층", "이름", "건수", "설명", "절"], layers, [6, 12, 8, 46, 8]),
        ("테스트케이스 및 결과", ["TC", "요구사항", "상태", "관측"], cases, [14, 22, 12, 88]),
    ]
    for i, (title, header, rows, widths) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        ws.append(header)
        for c in ws[1]:
            c.fill, c.font, c.alignment = head_fill, head_font, wrap
        for r in rows:
            ws.append(r)
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws = wb.create_sheet("읽어 주세요", 0)
    for line in [
        ["이 파일은 생성물입니다 — 손으로 고치지 마세요"],
        [""],
        ["원본", "docs/13_테스트계획.md"],
        ["다시 만들기", "make testplan-xlsx  (macOS · Linux)"],
        ["", "python scripts/build_testplan_xlsx.py --write  (Windows)"],
        [""],
        ["여기서 고친 내용은 다음 생성 때 전부 없어집니다."],
        [""],
        ["상태 표기", "✅ 통과 — 재현했고 관측값이 있다"],
        ["", "⚠️ 참고필요 — 자동화가 실제 경로가 아니거나, 일부 조건이 미검증"],
        ["", "⬜ 미실행 — 아직 아무도 안 돌렸다"],
        ["", "🔴 미적용 — 요구사항을 아직 못 채웠다"],
        [""],
        ["ID 안내", "TC 이름은 정본 요구사항 ID(FR-nn · NFR-nn)를 가리킵니다."],
        ["", "팀원 원본의 FR-AUTH-002 식 접두어와의 대응은 docs/10 §11."],
    ]:
        ws.append(line)
    ws["A1"].font = Font(bold=True, size=13, color="B03A2E")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 68

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


def main() -> int:
    ap = argparse.ArgumentParser(description="테스트계획 → 제출용 xlsx (생성물)")
    ap.add_argument("--write", action="store_true", help=f"{OUT.relative_to(ROOT)} 에 기록")
    args = ap.parse_args()

    if not DOC.exists():
        print(f"✗ {DOC.relative_to(ROOT)} 가 없다")
        return 1

    layers, cases = read_layers(), read_cases()
    print("━━ 테스트계획 → xlsx ━━")
    print(f"  층 {len(layers)} · 테스트케이스 {len(cases)}")

    # 🔴 표 형식이 바뀌면 조용히 빈 시트가 나간다. 여기서 크게 실패한다.
    if len(layers) != 5:
        print(f"  ✗ 층을 {len(layers)} 개 읽었다 — §1 의 다섯 줄 형식이 바뀌었는지 본다")
        return 1
    if not cases:
        print("  ✗ 테스트케이스를 하나도 못 읽었다 — §2.2 표 형식이 바뀌었는지 본다")
        return 1

    if args.write:
        _write(layers, cases)
        print(f"  ✓ {OUT.relative_to(ROOT)}")
    else:
        print("\n기록하려면 --write")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
