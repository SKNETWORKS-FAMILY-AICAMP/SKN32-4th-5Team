"""docs/10_요구사항정의서.md → 제출용 xlsx.

    python scripts/build_requirements_xlsx.py --write   # 제출/요구사항정의서.xlsx 기록
    make reqs-xlsx                                     # 같은 것 (Windows 에는 make 가 없다)
    python scripts/build_requirements_xlsx.py           # 미리 보기만

🔴 **xlsx 는 생성물이다. 손으로 고치지 않는다.** `build_rule_table.py` 와 같은 규약이다.

왜 이 스크립트가 있나 — 2026-08-27 에 요구사항 표가 `.xlsx` 로 **두 벌**(v3 · v4) 돌아다녔고,
두 판본 사이에서 **요구사항 한 줄(DB 백업)이 소리 없이 사라졌다.** `.xlsx` 는 diff 가 안 돼서
PR 로도 안 보이고, 두 판본이 우연히 남아 있어서 겨우 알았다.

사람이 고치는 것은 `.md` 하나뿐이고, 제출용 표는 여기서 나온다. 그러면
**갈라질 자리가 없다** (D-22).
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DOC = ROOT / "docs" / "10_요구사항정의서.md"
OUT = ROOT / "제출" / "요구사항정의서.xlsx"

#: 표 머리행이 이 값으로 시작하면 데이터가 아니다.
_HEADER_FIRST = {"ID", "FR", "항목", "#"}


def _cells(line: str) -> list[str]:
    return [c.strip() for c in line.strip().strip("|").split("|")]


def _is_rule(cells: list[str]) -> bool:
    return all(set(c) <= {"-", ":", ""} for c in cells)


def _table(text: str, heading_has: str, id_prefix: str) -> list[list[str]]:
    """제목에 `heading_has` 가 든 절에서 `id_prefix` 로 시작하는 행을 모은다."""
    rows: list[list[str]] = []
    heading = ""
    for line in text.splitlines():
        if line.startswith("#"):
            heading = line.lstrip("#").strip()
            continue
        if not line.lstrip().startswith("|"):
            continue
        cells = _cells(line)
        if not cells or _is_rule(cells) or cells[0] in _HEADER_FIRST:
            continue
        if heading_has in heading and cells[0].startswith(id_prefix):
            rows.append(cells)
    return rows


def _plain(cell: str) -> str:
    """마크다운 강조·코드 표시를 걷어 낸다. 상태 이모지는 남긴다."""
    cell = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", cell)  # 링크 → 글자만
    return re.sub(r"[*`~]", "", cell).strip()


def build() -> tuple[list[list[str]], list[list[str]], list[list[str]]]:
    text = DOC.read_text(encoding="utf-8")
    fr = [[_plain(c) for c in r] for r in _table(text, "기능 요구사항 (FR)", "FR-")]
    nfr = [[_plain(c) for c in r] for r in _table(text, "비기능", "NFR-")]
    trace = [[_plain(c) for c in r] for r in _table(text, "추적표", "FR-")]
    return fr, nfr, trace


def _write(fr, nfr, trace) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl 이 없다 — pip install -e '.[ingest]' -c constraints.txt")

    head_fill = PatternFill("solid", fgColor="3A4F37")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    sheets = [
        (
            "기능 요구사항",
            ["ID", "요구사항", "UC", "상태", "근거", "분류", "순위"],
            fr,
            [10, 62, 10, 8, 44, 14, 7],
        ),
        (
            "비기능 요구사항",
            ["ID", "요구사항", "측정 방법", "현재 실측", "근거"],
            nfr,
            [10, 56, 26, 26, 26],
        ),
        ("추적표", ["FR", "UC", "화면", "테스트", "상태"], trace, [10, 10, 18, 30, 8]),
    ]
    for i, (title, header, rows, widths) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        ws.append(header)
        for c in ws[1]:
            c.fill, c.font, c.alignment = head_fill, head_font, wrap
        for r in rows:
            ws.append((r + [""] * len(header))[: len(header)])
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # 🔴 이 시트가 없으면 다음 사람이 xlsx 를 직접 고친다.
    ws = wb.create_sheet("읽어 주세요", 0)
    for line in [
        ["이 파일은 생성물입니다 — 손으로 고치지 마세요"],
        [""],
        ["원본", "docs/10_요구사항정의서.md"],
        ["다시 만들기", "make reqs-xlsx  (macOS · Linux)"],
        ["", "python scripts/build_requirements_xlsx.py --write  (Windows)"],
        [""],
        ["여기서 고친 내용은 다음 생성 때 전부 없어집니다."],
        ["요구사항이 바뀌었으면 위 .md 를 고치고 다시 만드세요."],
        [""],
        ["상태 표기", "✅ 구현됐고 CI 가 검증한다"],
        ["", "🟡 구현됐으나 검증되지 않는다 (테스트가 없다)"],
        ["", "🔴 미구현"],
        ["", "⬜ 미착수"],
    ]:
        ws.append(line)
    ws["A1"].font = Font(bold=True, size=13, color="B03A2E")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


def main() -> int:
    ap = argparse.ArgumentParser(description="요구사항정의서 → 제출용 xlsx (생성물)")
    ap.add_argument("--write", action="store_true", help=f"{OUT.relative_to(ROOT)} 에 기록")
    args = ap.parse_args()

    if not DOC.exists():
        print(f"✗ {DOC.relative_to(ROOT)} 가 없다")
        return 1

    fr, nfr, trace = build()
    print("━━ 요구사항정의서 → xlsx ━━")
    print(f"  기능 요구사항 {len(fr)} · 비기능 {len(nfr)} · 추적 {len(trace)}")

    if not fr or not nfr:
        print("  ✗ 표를 못 읽었다 — docs/10 의 표 형식이 바뀌었는지 본다")
        return 1
    if len(fr) != len(trace):
        print(f"  ✗ FR {len(fr)} 과 추적 {len(trace)} 이 다르다 —")
        print("    check_requirements.py 를 먼저 돌린다")
        return 1

    if args.write:
        _write(fr, nfr, trace)
        print(f"  ✓ {OUT.relative_to(ROOT)}")
    else:
        print("\n기록하려면 --write")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
